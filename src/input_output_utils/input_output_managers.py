
import os
from openpyxl.styles import PatternFill
from typing import Union
from enum import Enum

import pandas as pd


def excel_only(func):
    def wrapper(self, *args, **kwargs):
        if not isinstance(self._manager, _ExcelManager):
            raise TypeError(f"{func.__name__} can only be used for Excel output files.\n"
                            f"Output path {self._output_path} is not an Excel path.")
        return func(self, *args, **kwargs)
    return wrapper


class HexColor(Enum):
    DARK_BLUE = 'ced8ff'
    BLUE = 'd8e0ff'
    LIGHT_BLUE = 'f1f4ff'
    DARK_GREEN = '2da15d'


class ExcelFormat:

    def __init__(self, color: HexColor, font_size: int, is_bold: bool):
        self.color = color
        self.font_size = font_size
        self.is_bold = is_bold


class CellFormatMap:

    def __init__(self):
        self._cell_map = dict()

    def format_cell(self, row_idx: int, col_idx: int, excel_format: ExcelFormat, exist_ok=False):
        t = (row_idx, col_idx)
        if t in self._cell_map and not exist_ok:
            raise KeyError(f"Cell at (row, col) ({row_idx}, {col_idx}) already exists in {self.__class__.__name__}.")

        self._cell_map[(row_idx, col_idx)] = excel_format

    def iter_cells(self):
        for row_idx, col_idx, excel_format in self._cell_map.items():
            yield row_idx, col_idx, excel_format


class DataTable:

    def __init__(self,
                 df: pd.DataFrame):
        self.df = df

        self.format_map = CellFormatMap()

    @property
    def total_rows(self):
        return self.df.shape[0]

    @property
    def total_columns(self):
        return self.df.shape[1]

    def insert_empty_rows(self, how_many: int = 1):
        for _ in range(how_many):
            self.df.loc[len(self.df)] = None

    def insert_empty_columns(self, how_many: int = 1):
        num_rows = self.total_rows
        for _ in range(how_many):
            self.df.insert(self.total_columns, None, [None for _ in range(num_rows)])

    def format_row(self, row_idx: int, excel_format: ExcelFormat):
        for col_idx in range(self.total_columns):
            self.format_map.format_cell(row_idx=row_idx, col_idx=col_idx, excel_format=excel_format)

    def format_column(self, col_idx: int, excel_format: ExcelFormat):
        for row_idx in range(self.total_rows):
            self.format_map.format_cell(row_idx=row_idx, col_idx=col_idx, excel_format=excel_format)


class DataSheet:

    def __init__(self):
        self._tables = []

    @property
    def shape(self) -> tuple:
        rows = 0
        cols = 0

        for (row_start_idx, col_start_idx), table in self._tables:
            rows = max(
                rows,
                table.total_rows + row_start_idx
            )
            cols = max(
                cols,
                table.total_columns + col_start_idx
            )

        return rows, cols

    def insert_data_table(self, data_table: DataTable, row_start_idx: int, col_start_idx: int):
        self._tables.append(
            ((row_start_idx, col_start_idx), data_table)
        )


    def create_format_map(self) -> CellFormatMap:
        master_map = CellFormatMap()
        for (row_start_idx, col_start_idx), table in self._tables:
            table_map = table.format_map
            for (row_idx, col_idx), excel_format in table_map._cell_map.items():
                master_map.format_cell(
                    row_idx=row_start_idx + row_idx,
                    col_idx=col_start_idx + col_idx,
                    excel_format=excel_format
                )
        return master_map

    def iter_tables(self):
        for (row_start_idx, col_start_idx), table in self._tables:
            yield row_start_idx, col_start_idx, table

class _ExcelManager:

    def __init__(self, output_path: str):
        self._writer = pd.ExcelWriter(output_path, engine='openpyxl')
        self._sheet_data = dict()

    def add_data_sheet(self, data_sheet: DataSheet, sheet_name: str):
        self._sheet_data[sheet_name] = data

        data.to_excel(self._writer, sheet_name=sheet_name, index=False)

        if color_map:
            ws = self._fetch_worksheet(sheet_name)
            for (cell_x, cell_y), hex_color in color_map._cell_map:
                if hex_color in self._pattern_fills:
                    fill = self._pattern_fills[hex_color]
                else:
                    fill = PatternFill(start_color=hex_color, fill_type='solid')
                    self._pattern_fills[hex_color] = fill

                ws.cell(cell_x, cell_y).fill = fill

    def _fetch_worksheet(self, sheet_name: str):
        return self._writer.sheets[sheet_name]

    def autofit_column_widths(self, sheet_name: str=None):
        """

        :param sheet_name: If provided, only autofits column widths for the singular sheet. Otherwise, autofit column widths for all sheets.
        :return:
        """
        from openpyxl.utils import get_column_letter

        sheet_names = [sheet_name] if sheet_name else list(self._sheet_data.keys())

        for sheet_name in sheet_names:
            worksheet = self._fetch_worksheet(sheet_name=sheet_name)
            data = self._sheet_data[sheet_name]

            for i in range(data.shape[1]):
                col_series = data.iloc[:, i]
                col_name = data.columns[i]
                max_len = max(col_series.astype(str).map(len).max(), len(str(col_name)))
                worksheet.column_dimensions[get_column_letter(i + 1)].width = max_len + 2

    def export(self):
        self._writer.close()



class _CsvManager:

    def __init__(self,
                 output_path: str,
                 data: pd.DataFrame = None):
        self._output_path = output_path
        self._data = data

    def add_data(self, data: pd.DataFrame):
        print("Overwriting previous data in CsvManager.")
        self._data = data

    def export(self):
        if not self._data:
            raise ValueError(f"CsvManager has no data to export.")

        self._data.to_csv(self._output_path)


class Sheet:

    def __init__(self,
                 df: pd.DataFrame,
                 sheet_name: str):
        self.df = df
        self.sheet_name = sheet_name

class OutputManager:
    _manager: Union[_ExcelManager, _CsvManager]

    def __init__(self,
                 output_path: str,
                 data: pd.DataFrame = None,
                 sheet_name: str = None):
        self.output_path = output_path

        ext = os.path.splitext(output_path)[1].lower()
        if ext in ('.xlsx', '.xls'):
            self._manager = _ExcelManager(output_path)
        elif ext == '.csv':
            self._manager = _CsvManager(output_path)
        else:
            raise ValueError(f"Output path {output_path} is not supported.")

        if data:
            self.add_data(
                data=data,
                sheet_name=sheet_name
            )

        return self

    @property
    def is_excel(self):
        return isinstance(self._manager, _ExcelManager)

    @property
    def is_csv(self):
        return isinstance(self._manager, _CsvManager)

    @excel_only
    def autofit_column_widths(self, sheet_name: str = None):
        self._manager.autofit_column_widths(sheet_name=sheet_name)

    def add_data(self,
                 sheets: list[Sheet]):
        if self.is_excel:
            if isinstance(data, pd.DataFrame) and not sheet_name:
                raise ValueError(f"Sheet name must be included as an argument when the a singular DataFrame is supplied "
                                 f"and the output is Excel.")

            if isinstance(data, dict):
                non_dataframes = [v for v in data.values() if not isinstance(v, pd.DataFrame)]
                if non_dataframes:
                    raise TypeError(f"Non-DataFrames supplied as values in dict when calling add_data.\n"
                                    f"Invalid types: {[type(non_df) for non_df in non_dataframes]}")

                for sheet in sheets:
                    self._manager.add_data(
                        data=sheet.data,
                        sheet_name=sheet.sheet_name,
                        color_map=sheet.cell_color_map
                    )

            return self

        if self.is_csv:
            if isinstance(sheets, list):
                raise TypeError(f"Supplying a multiple Sheets is invalid for CSV.")

            self._manager.add_data(
                data=data
            )

            return self

    def export(self):
        self._manager.export()

        return self
